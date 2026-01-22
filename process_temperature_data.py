#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
温度数据处理脚本
功能：
1. 解析data/data1.txt中的温度数据
2. 根据template/template.xlsx中的通道位置映射
3. 生成包含均值图和所有测试块数据的新Excel文件
4. 应用红绿渐变条件格式
"""

import re
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule
from collections import defaultdict
from typing import List, Dict, Tuple, Optional


def parse_data_file(file_path: str) -> Tuple[List[Dict[int, float]], List[str]]:
    """
    解析data1.txt文件，按#####数字#####分块，提取valid=1的chnl和temp数据
    
    Returns:
        Tuple[List[Dict[int, float]], List[str]]: 
            - 每个测试块的 {chnl: temp} 字典列表
            - 每个测试块的标题列表（分隔符中的数字）
    """
    data_path = Path(file_path)
    if not data_path.exists():
        raise FileNotFoundError(f"数据文件不存在: {file_path}")
    
    blocks = []
    block_titles = []
    current_block = {}
    current_title = None
    
    with open(data_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            
            # 检测新的测试块分隔符
            match_separator = re.match(r'^#####(\d+)#####$', line)
            if match_separator:
                # 如果当前块有数据，保存它
                if current_block:
                    blocks.append(current_block)
                    block_titles.append(current_title if current_title else "Unknown")
                current_block = {}
                current_title = match_separator.group(1)  # 提取分隔符中的数字
                continue
            
            # 解析数据行: chnl X, valid Y, temp Z
            match = re.match(r'chnl\s+(\d+),\s+valid\s+(\d+),\s+temp\s+([-\d.]+)', line)
            if match:
                chnl = int(match.group(1))
                valid = int(match.group(2))
                temp = float(match.group(3))
                
                # 只保存valid=1的数据
                if valid == 1:
                    current_block[chnl] = temp
    
    # 保存最后一个块
    if current_block:
        blocks.append(current_block)
        block_titles.append(current_title if current_title else "Unknown")
    
    return blocks, block_titles


def read_template_mapping(template_path: str) -> Tuple[Dict[Tuple[int, int], int], int, int]:
    """
    读取模板Excel，建立通道号到行列位置的映射关系
    
    Returns:
        Tuple[Dict[Tuple[int, int], int], int, int]:
            - 映射字典: {(row, col): chnl}
            - 最大行数
            - 最大列数
    """
    template_file = Path(template_path)
    if not template_file.exists():
        raise FileNotFoundError(f"模板文件不存在: {template_path}")
    
    wb = load_workbook(template_file)
    ws = wb['sheet1']
    
    mapping = {}  # {(row, col): chnl}
    max_row = 0
    max_col = 0
    
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell.value is not None:
                try:
                    chnl = int(cell.value)
                    mapping[(row_idx, col_idx)] = chnl
                    max_row = max(max_row, row_idx)
                    max_col = max(max_col, col_idx)
                except (ValueError, TypeError):
                    # 如果不是整数，忽略
                    pass
    
    return mapping, max_row, max_col


def calculate_average_temps(blocks: List[Dict[int, float]]) -> Dict[int, float]:
    """
    计算所有测试块中每个通道的平均温度
    
    Args:
        blocks: 所有测试块的数据列表
        
    Returns:
        Dict[int, float]: {chnl: average_temp}
    """
    # 收集每个通道的所有有效温度值
    chnl_temps = defaultdict(list)
    
    for block in blocks:
        for chnl, temp in block.items():
            chnl_temps[chnl].append(temp)
    
    # 计算平均值
    avg_temps = {}
    for chnl, temps in chnl_temps.items():
        avg_temps[chnl] = sum(temps) / len(temps)
    
    return avg_temps


def write_title(ws, title: str, row: int, max_col: int):
    """
    在指定行写入标题，合并单元格并设置样式
    
    Args:
        ws: Excel工作表对象
        title: 标题文本
        row: 行号
        max_col: 最大列号（用于合并单元格）
    """
    title_cell = ws.cell(row=row, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    if max_col > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)


def write_block_to_excel(ws, block_data: Dict[int, float], 
                        mapping: Dict[Tuple[int, int], int],
                        start_row: int) -> int:
    """
    将一个测试块的数据写入Excel的指定起始行
    
    Args:
        ws: Excel工作表对象
        block_data: 当前块的数据 {chnl: temp}
        mapping: 位置映射 {(row, col): chnl}
        start_row: 起始行号
        
    Returns:
        int: 下一个块的起始行号（当前块结束行 + 2，留一个空行）
    """
    # 找到当前块需要的最大行数
    max_template_row = max(row for row, _ in mapping.keys())
    
    # 写入数据
    for (template_row, template_col), chnl in mapping.items():
        excel_row = start_row + template_row - 1  # 转换为Excel中的实际行号
        excel_col = template_col
        
        # 如果当前块有这个通道的有效数据，写入温度值
        if chnl in block_data:
            ws.cell(row=excel_row, column=excel_col, value=block_data[chnl])
    
    # 返回下一个块的起始行（当前块结束行 + 2，留一个空行）
    return start_row + max_template_row + 1


def apply_color_scale(ws, min_temp: float, max_temp: float):
    """
    对有值的单元格应用红绿渐变条件格式
    红色=高温，绿色=低温
    
    Args:
        ws: Excel工作表对象
        min_temp: 最小温度值
        max_temp: 最大温度值
    """
    # 获取工作表的所有有数据的范围
    # 找到所有有值的单元格范围
    max_row = ws.max_row
    max_col = ws.max_column
    
    if max_row == 0 or max_col == 0:
        return
    
    # 创建颜色渐变规则
    # 使用三色渐变：绿色(低温) -> 黄色(中温) -> 红色(高温)
    color_scale = ColorScaleRule(
        start_type='num',
        start_value=min_temp,
        start_color='00FF00',  # 绿色（低温）
        mid_type='num',
        mid_value=(min_temp + max_temp) / 2,
        mid_color='FFFF00',    # 黄色（中温）
        end_type='num',
        end_value=max_temp,
        end_color='FF0000'     # 红色（高温）
    )
    
    # 应用到整个数据区域
    data_range = f'A1:{ws.cell(row=max_row, column=max_col).coordinate}'
    ws.conditional_formatting.add(data_range, color_scale)


def main():
    """主函数"""
    # 文件路径
    data_file = 'data/data1.txt'
    template_file = 'template/template.xlsx'
    # 结果文件放在 result 目录下
    output_dir = Path('result')
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / 'result.xlsx'
    
    print("开始处理温度数据...")
    
    # 1. 解析数据文件
    print("1. 解析数据文件...")
    blocks, block_titles = parse_data_file(data_file)
    print(f"   找到 {len(blocks)} 个测试块")
    
    # 2. 读取模板映射
    print("2. 读取模板映射...")
    mapping, max_template_row, max_template_col = read_template_mapping(template_file)
    print(f"   模板大小: {max_template_row} 行 x {max_template_col} 列")
    print(f"   找到 {len(mapping)} 个通道位置")
    
    # 3. 计算均值
    print("3. 计算平均温度...")
    avg_temps = calculate_average_temps(blocks)
    print(f"   计算了 {len(avg_temps)} 个通道的平均值")
    
    # 4. 创建新Excel
    print("4. 生成Excel文件...")
    wb = Workbook()
    ws = wb.active
    ws.title = "result"
    
    # 收集所有温度值，用于条件格式
    all_temps = []
    
    # 4.1 先写入均值图标题和数据（最上面）
    print("   写入均值数据图...")
    # 写入均值图标题
    write_title(ws, "Average Temperature Map", row=1, max_col=max_template_col)
    
    # 写入均值图数据（从第2行开始）
    current_row = 2
    for (template_row, template_col), chnl in mapping.items():
        excel_row = current_row + template_row - 1
        excel_col = template_col
        if chnl in avg_temps:
            temp = avg_temps[chnl]
            ws.cell(row=excel_row, column=excel_col, value=temp)
            all_temps.append(temp)
    
    # 4.2 写入各个测试块（均值图下方，留空行）
    # 均值图结束行 = current_row + max_template_row - 1
    # 下一个位置 = 均值图结束行 + 2（留一个空行）
    current_row = current_row + max_template_row + 1
    
    for i, (block, title) in enumerate(zip(blocks, block_titles), 1):
        print(f"   写入测试块 {i} (标题: {title})...")
        
        # 写入块标题
        write_title(ws, f"Block {i} (#####{title}#####)", row=current_row, max_col=max_template_col)
        
        # 写入块数据（从标题下一行开始）
        data_start_row = current_row + 1
        block_end_row = write_block_to_excel(ws, block, mapping, data_start_row)
        
        # 收集温度值
        for temp in block.values():
            all_temps.append(temp)
        
        # 下一个块从当前块结束行 + 1个空行开始
        current_row = block_end_row + 1
    
    # 5. 应用条件格式
    if all_temps:
        print("5. 应用条件格式...")
        min_temp = min(all_temps)
        max_temp = max(all_temps)
        print(f"   温度范围: {min_temp} ~ {max_temp}")
        apply_color_scale(ws, min_temp, max_temp)
    
    # 6. 保存文件
    print(f"6. 保存文件到: {output_file}")
    wb.save(str(output_file))
    
    print("处理完成！")


if __name__ == '__main__':
    main()
